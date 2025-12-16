import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog
import os
from colorama import Fore, Style

def select_file(title):
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    f = filedialog.askopenfilename(title=title, filetypes=[("Excel/CSV", "*.xlsx *.csv")])
    root.destroy()
    return f

def _read_sdt_sheet(path, sheet_name):
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, header=0, dtype=str)
        if len(df) > 2:
            df = df.iloc[2:].reset_index(drop=True)
        else:
            return pd.DataFrame(columns=df.columns)
        df.columns = [str(c).strip().upper() for c in df.columns]
        return df
    except Exception as e:
        print(f"{Fore.RED}Error reading sheet {sheet_name}: {e}{Style.RESET_ALL}")
        return pd.DataFrame()

def _normalize_legacy_cols(cols):
    mapping = {}
    for c in cols:
        clean = str(c).strip().upper()
        if len(clean) == 6 and clean[:2].isalpha():
            short = clean[2:]
            mapping[short] = clean
        else:
            mapping[clean] = clean
    return mapping

def load_and_join(legacy_path, target_path):
    # 1. Load Legacy
    print("\n   -> analyzing Legacy file...")
    legacy_sheet_name = "CSV_Data"
    
    if legacy_path.endswith('.csv'): 
        df_leg = pd.read_csv(legacy_path, dtype=str)
    else: 
        # Ask for sheet
        wb_leg = openpyxl.load_workbook(legacy_path, read_only=True)
        sheets_leg = wb_leg.sheetnames
        wb_leg.close()
        
        if len(sheets_leg) == 1:
            legacy_sheet_name = sheets_leg[0]
        else:
            print(f"\n   Available Legacy Sheets:")
            for i, s in enumerate(sheets_leg): print(f"   {i+1}. {s}")
            while True:
                try:
                    val = input("\n   Select LEGACY Sheet [Number]: ").strip()
                    idx = int(val) - 1
                    if 0 <= idx < len(sheets_leg):
                        legacy_sheet_name = sheets_leg[idx]
                        break
                except ValueError: pass
        
        print(f"   -> Loading Legacy Sheet: {legacy_sheet_name}")
        df_leg = pd.read_excel(legacy_path, sheet_name=legacy_sheet_name, dtype=str)
    
    df_leg.columns = [str(c).strip().upper() for c in df_leg.columns]
    
    # 2. Analyze Target (SDT)
    print("   -> Analyzing Target (SDT) file structure...")
    wb = openpyxl.load_workbook(target_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()

    print("\n   Available Target Sheets:")
    for i, s in enumerate(sheets): print(f"   {i+1}. {s}")
    
    # 3. Select Target Sheets
    main_sheet = None
    while True:
        try:
            val = input("\n   Select MAIN Target Sheet (Primary Key holder) [Number]: ").strip()
            idx = int(val) - 1
            if 0 <= idx < len(sheets):
                main_sheet = sheets[idx]; break
        except ValueError: pass

    print("\n   Enter other Target sheets to merge (comma separated numbers, or Enter for none):")
    other_input = input("   >> ").strip()
    other_sheets = []
    if other_input:
        for x in other_input.split(','):
            try:
                idx = int(x.strip()) - 1
                if 0 <= idx < len(sheets) and sheets[idx] != main_sheet: other_sheets.append(sheets[idx])
            except: pass

    # 4. Load & Stitch Target
    print(f"\n   -> Loading Main Sheet: {main_sheet}...")
    df_tgt = _read_sdt_sheet(target_path, main_sheet)
    
    for s in other_sheets:
        print(f"   -> Merging Sheet: {s}...")
        df_other = _read_sdt_sheet(target_path, s)
        common = list(set(df_tgt.columns) & set(df_other.columns))
        keys = [k for k in ['CONO', 'DIVI', 'ITNO', 'CUNO', 'SUNO', 'FACI', 'WHLO'] if k in common]
        if keys: df_tgt = pd.merge(df_tgt, df_other, on=keys, how='left', suffixes=('', f'_{s}'))

    # 5. Join Legacy <-> Target
    print(f"\n   -> Matching Legacy to Target...")
    leg_map_rev = _normalize_legacy_cols(df_leg.columns)
    
    join_key_tgt = None
    join_key_leg = None

    for key in ['ITNO', 'CUNO', 'SUNO', 'CONO', 'ORDN']:
        if key in df_tgt.columns and key in leg_map_rev:
            join_key_tgt = key
            join_key_leg = leg_map_rev[key]
            break
            
    if not join_key_tgt:
        raise ValueError("Could not find a common Key (ITNO, CUNO, etc.) between files.")

    print(f"      Joining on: Legacy[{join_key_leg}] == Target[{join_key_tgt}]")

    df_leg['__KEY__'] = df_leg[join_key_leg].astype(str).str.strip()
    df_tgt['__KEY__'] = df_tgt[join_key_tgt].astype(str).str.strip()

    df_combined = pd.merge(df_leg, df_tgt, on='__KEY__', how='inner', suffixes=('_SRC', '_TGT'))
    
    src_cols = [c + '_SRC' if c in df_tgt.columns else c for c in df_leg.columns if c != '__KEY__']
    tgt_cols = [c + '_TGT' if c in df_leg.columns else c for c in df_tgt.columns if c != '__KEY__']
    
    valid_src = [c for c in src_cols if c in df_combined.columns]
    valid_tgt = [c for c in tgt_cols if c in df_combined.columns]

    # Return legacy_sheet_name too!
    return df_combined, valid_src, valid_tgt, legacy_sheet_name